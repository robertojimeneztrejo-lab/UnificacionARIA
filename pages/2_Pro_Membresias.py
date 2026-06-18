import streamlit as st
import google.generativeai as genai
import json
import re
import io
import requests
import pandas as pd
import urllib.parse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

st.set_page_config(
    page_title="ARIA Membresías",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Libre+Baskerville:wght@400;700&family=Source+Sans+3:wght@300;400;500;600&display=swap');

:root {
    --navy:        #0D2B4E;
    --navy-mid:    #1A3F6F;
    --navy-light:  #2E5FA3;
    --gold:        #C8973A;
    --gold-light:  #E8B84B;
    --surface:     #F4F6F9;
    --white:       #FFFFFF;
    --text:        #1A1F2E;
    --text-soft:   #3D4A5C;
    --muted:       #6B7A8D;
    --border:      #C8D0DC;
    --border-light:#DDE3EC;
    --green:       #1A5232;
    --green-bg:    #E8F5EE;
    --red:         #7B1E1E;
    --red-bg:      #FBEAEA;
}

html, body, [data-testid="stAppViewContainer"] {
    background-color: var(--surface) !important;
    font-family: 'Source Sans 3', sans-serif;
    color: var(--text);
}

[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0D2B4E 0%, #0A2240 100%) !important;
    border-right: 3px solid var(--gold) !important;
}
[data-testid="stSidebar"] * { color: #C8D8EC !important; }
[data-testid="stSidebar"] .stMarkdown h2 {
    font-family: 'Libre Baskerville', serif !important;
    color: #FFFFFF !important;
    font-size: 1rem !important;
    letter-spacing: 0.05em !important;
}
[data-testid="stSidebar"] .stMarkdown h3 {
    font-family: 'Source Sans 3', sans-serif !important;
    color: var(--gold-light) !important;
    font-size: 0.72rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.12em !important;
    text-transform: uppercase !important;
}
[data-testid="stSidebar"] hr { border-color: #1E4070 !important; }
.stCheckbox label { font-size: 0.82rem !important; font-family: 'Source Sans 3', sans-serif !important; }
[data-testid="stSidebar"] .stCheckbox label,
[data-testid="stSidebar"] .stCheckbox label p { color: #A8BECE !important; }
[data-testid="stSidebar"] .stRadio label,
[data-testid="stSidebar"] .stRadio label p { color: #A8BECE !important; }

.inst-header {
    background: linear-gradient(135deg, #0D2B4E 0%, #1A3F6F 60%, #2E5FA3 100%);
    border-radius: 10px;
    padding: 1.6rem 2rem;
    margin-bottom: 1.5rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    border-bottom: 4px solid var(--gold);
    box-shadow: 0 4px 20px rgba(13,43,78,0.18);
}
.inst-header-left { display: flex; align-items: center; gap: 1.2rem; }
.inst-logo {
    width: 52px; height: 52px;
    background: rgba(255,255,255,0.08);
    border: 2px solid var(--gold);
    border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
    font-size: 1.6rem; flex-shrink: 0;
}
.inst-title {
    font-family: 'Libre Baskerville', serif;
    font-size: 1.7rem; font-weight: 700;
    color: #FFFFFF; letter-spacing: 0.04em; line-height: 1.1;
}
.inst-title span { color: var(--gold-light); }
.inst-subtitle {
    font-size: 0.75rem; color: #8BAAC8;
    letter-spacing: 0.08em; text-transform: uppercase; margin-top: 4px;
}
.inst-badge {
    background: rgba(200,151,58,0.12);
    border: 1px solid rgba(200,151,58,0.4);
    border-radius: 6px; padding: 0.5rem 1rem; text-align: right;
}
.inst-badge-label { font-size: 0.62rem; color: #8BAAC8; text-transform: uppercase; letter-spacing: 0.1em; }
.inst-badge-value { font-size: 0.82rem; color: var(--gold-light); font-weight: 600; margin-top: 2px; }

.inst-mascot-box {
    display: flex; flex-direction: column; align-items: center;
    gap: 4px; flex-shrink: 0;
}
.inst-mascot-label {
    font-size: 0.62rem; color: #8BAAC8;
    text-transform: uppercase; letter-spacing: 0.08em;
    text-align: center;
}
.mascot-searching .mag-glass { animation: mag-sweep 1.4s ease-in-out infinite; }
.mascot-searching .mascot-body { animation: mascot-bob 1.4s ease-in-out infinite; }
.mascot-idle .mag-glass { animation: none; }
.mascot-idle .mascot-body { animation: none; }
@keyframes mag-sweep {
    0%   { transform: rotate(-8deg) translateX(0px); }
    50%  { transform: rotate(8deg) translateX(3px); }
    100% { transform: rotate(-8deg) translateX(0px); }
}
@keyframes mascot-bob {
    0%   { transform: translateY(0px); }
    50%  { transform: translateY(-2px); }
    100% { transform: translateY(0px); }
}

.stTextInput input {
    border: 1.5px solid var(--border) !important;
    border-radius: 6px !important;
    font-family: 'Source Sans 3', sans-serif !important;
    font-size: 0.95rem !important;
    background: var(--white) !important;
    color: var(--text) !important;
}
.stTextInput input:focus {
    border-color: var(--navy-light) !important;
    box-shadow: 0 0 0 3px rgba(46,95,163,0.12) !important;
}

.stButton > button {
    background: var(--navy) !important;
    color: #FFFFFF !important;
    font-family: 'Source Sans 3', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.85rem !important;
    border: 1.5px solid var(--navy-light) !important;
    border-radius: 6px !important;
    transition: all 0.15s !important;
}
.stButton > button:hover {
    background: var(--navy-light) !important;
    border-color: var(--gold) !important;
}

.status-bar {
    background: var(--navy);
    color: var(--gold-light);
    font-family: 'Source Sans 3', sans-serif;
    font-size: 0.78rem; font-weight: 600;
    padding: 0.45rem 1rem;
    border-radius: 5px;
    border-left: 3px solid var(--gold);
    letter-spacing: 0.06em;
    margin-bottom: 1rem;
    display: inline-block;
}

.result-card {
    background: var(--white);
    border: 1px solid var(--border-light);
    border-radius: 8px;
    padding: 1.2rem 1.5rem;
    margin-bottom: 0.9rem;
    position: relative; overflow: hidden;
    box-shadow: 0 1px 4px rgba(13,43,78,0.07);
    transition: box-shadow 0.2s, border-color 0.2s;
}
.result-card::before {
    content: '';
    position: absolute; top: 0; left: 0;
    width: 4px; height: 100%;
    background: linear-gradient(180deg, var(--gold) 0%, var(--navy-light) 100%);
}
.result-card:hover {
    box-shadow: 0 4px 16px rgba(13,43,78,0.13);
    border-color: var(--navy-light);
}
.result-card.approved::before { background: var(--green); }
.result-card.rejected::before { background: var(--red); }
.result-card.rejected { opacity: 0.6; }

.card-title {
    font-family: 'Libre Baskerville', serif;
    font-size: 1rem; font-weight: 700; color: var(--navy); margin-bottom: 0.15rem;
}
.card-url { font-size: 0.77rem; color: var(--muted); margin-bottom: 0.7rem; word-break: break-all; }
.card-url a { color: var(--navy-light); text-decoration: none; }
.card-url a:hover { text-decoration: underline; }
.card-section-title {
    font-size: 0.65rem; font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.1em;
    color: var(--muted); margin-bottom: 0.2rem; margin-top: 0.65rem;
    border-bottom: 1px solid var(--border-light); padding-bottom: 2px;
}
.card-text { font-size: 0.87rem; color: var(--text-soft); line-height: 1.55; }

.badge { font-size: 0.7rem; font-weight: 600; padding: 2px 9px; border-radius: 3px; font-family: 'Source Sans 3', sans-serif; }
.badge-type-institucional { background: #E8F0FA; color: #1A3F6F; border: 1px solid #B8CEE8; }
.badge-type-corporativa   { background: #FDF3E3; color: #7D4E0F; border: 1px solid #E8C88A; }
.badge-type-academica     { background: #EAF3F7; color: #2D5A6E; border: 1px solid #B8DCE8; }
.badge-type-estudiantil   { background: #E8F5EE; color: #1A5232; border: 1px solid #A9D6BC; }
.badge-type-individual    { background: #F2EEF9; color: #4A2D8A; border: 1px solid #C5B3E8; }
.badge-type-asociado      { background: #FDEDF0; color: #8A1F3D; border: 1px solid #E8AFC0; }
.badge-region { background: #F0F3F8; color: #2D3D52; border: 1px solid #C0CBDA; }
.badge-access { background: #E8F5EE; color: #1A5232; border: 1px solid #A9D6BC; }
.badge-priority { background: var(--gold); color: #FFFFFF; border: 1px solid var(--gold-light); }

.stars       { font-size: 0.95rem; color: var(--gold); letter-spacing: 1px; }
.stars-empty { font-size: 0.95rem; color: var(--border); letter-spacing: 1px; }

.verified-badge { display: inline-block; font-size: 0.65rem; padding: 2px 7px; border-radius: 3px; margin-left: 8px; vertical-align: middle; font-weight: 600; }
.verified-ok  { background: var(--green-bg); color: var(--green); border: 1px solid #A9D6BC; }
.verified-no  { background: var(--red-bg); color: var(--red); border: 1px solid #E8AAAA; }
.verified-unk { background: #F0F3F8; color: #3D4A5C; border: 1px solid #C0CBDA; }

.empty-state {
    text-align: center; padding: 3.5rem 1rem;
    color: var(--muted); font-size: 0.95rem;
    background: var(--white);
    border: 1px dashed var(--border);
    border-radius: 8px; margin-top: 1rem;
}
.empty-icon { font-size: 2.2rem; margin-bottom: 0.8rem; display: block; opacity: 0.5; }

[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stNumberInput label { color: #A8BECE !important; font-size: 0.82rem !important; }

div[data-testid="stRadio"] > label { display: none; }

.contact-input input {
    border: 1.5px solid var(--border-light) !important;
    border-radius: 5px !important;
    font-size: 0.82rem !important;
    padding: 0.35rem 0.6rem !important;
}
</style>
""", unsafe_allow_html=True)

# ── API Key desde Secrets ─────────────────────────────────────────────────────
api_key = st.secrets.get("GEMINI_API_KEY", "")

# ── Google Sheets URLs ────────────────────────────────────────────────────────
SHEETS_CSV_URL  = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTmA6CzuX4HnRv_gfKHdZB4GcezxNGq0g5nUY7OkFyEbPeYfkSbMgeSg7O20WoqPs-YRVF0qVc3AdRA/pub?output=csv"
APPS_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbxWfJrZB1Q1IpBaH0nLOebv5C6PFZRm2BqChM-EQRS6xjgsBYnuE7A82hgQCi-0kdHiMA/exec"


@st.cache_data(ttl=60)
def load_existing_memberships():
    """Lee membresías ya obtenidas desde Google Sheets (refresca cada 60 seg)."""
    try:
        df = pd.read_csv(SHEETS_CSV_URL, header=None)
        nombres = df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
        if nombres and nombres[0].lower() in ["nombre", "name", "membresía", "membresia", "plataforma"]:
            nombres = nombres[1:]
        return [n for n in nombres if n]
    except Exception:
        return []


def save_to_sheets(results, decisions):
    """Envía los resultados (con su decisión aprobado/rechazado y correo) al Apps Script."""
    try:
        payload = []
        for idx, r in enumerate(results):
            nombre = r.get("nombre", "").strip()
            if not nombre:
                continue
            decision = decisions.get(idx, {})
            estatus = decision.get("estatus", "Pendiente")
            correo = decision.get("correo", "").strip() if estatus == "Aprobado" else ""

            payload.append({
                "nombre": nombre,
                "url":    r.get("url", "").strip(),
                "region": r.get("region", "").strip(),
                "tipo":   r.get("tipo_membresia", "").strip(),
                "tema":   (st.session_state.get("last_topic", "") or "General").strip(),
                "estatus": estatus,
                "correo": correo
            })

        if not payload:
            return False, "Sin resultados para guardar"

        nombres_encoded = urllib.parse.quote(json.dumps(payload, ensure_ascii=False))
        url = f"{APPS_SCRIPT_URL}?nombres={nombres_encoded}"

        resp = requests.get(url, timeout=25, allow_redirects=True)
        raw = resp.text.strip()

        if not raw:
            return False, "Respuesta vacía del servidor"
        if raw.startswith("<!"):
            return False, "El Apps Script devolvió HTML. Verifica que esté desplegado con acceso 'Cualquier persona'."

        data = json.loads(raw)
        if data.get("status") == "ok":
            return True, data
        return False, data.get("message", "Error desconocido")

    except json.JSONDecodeError:
        return False, f"Respuesta no JSON: {raw[:200]}"
    except Exception as ex:
        return False, str(ex)


# ── Configuración fija ────────────────────────────────────────────────────────
MAX_RESULTS = 5

# ── Palabras clave de prioridad ──────────────────────────────────────────────
KEYWORD_OPTIONS = ["Association", "League", "Alliance", "Society", "Charter",
                    "Royal College", "Organization"]

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🎓 ARIA\n### Filtros de Búsqueda")
    st.markdown("---")

    # Indicador de lista de exclusión
    excluidas_preview = load_existing_memberships()
    n_exc = len(excluidas_preview)
    if n_exc > 0:
        st.markdown(
            f'''<div style="background:#0A2240;border:1px solid #C8973A;border-radius:8px;padding:0.6rem 0.8rem;margin-bottom:0.5rem;">
            <div style="font-size:0.65rem;letter-spacing:0.1em;text-transform:uppercase;color:#E8B84B;font-weight:700;margin-bottom:3px;">📋 Lista de exclusión</div>
            <div style="font-size:0.82rem;color:#A8BECE;"><strong style="color:#E8B84B;">{n_exc}</strong> membresías ya registradas serán excluidas</div>
            </div>''',
            unsafe_allow_html=True
        )
    else:
        st.markdown(
            '<div style="background:#0A2240;border:1px solid #1E4070;border-radius:8px;padding:0.6rem 0.8rem;margin-bottom:0.5rem;">'
            '<div style="font-size:0.65rem;letter-spacing:0.1em;text-transform:uppercase;color:#6B7A8D;font-weight:700;margin-bottom:3px;">📋 Lista de exclusión</div>'
            '<div style="font-size:0.82rem;color:#6B7A8D;">Sin datos o cargando...</div></div>',
            unsafe_allow_html=True
        )
    st.markdown("---")

    st.markdown("### 🌐 Modo de búsqueda")
    modo_busqueda = st.radio(
        "Modo",
        options=["🔴 Live Crawl (Google Search)", "🧠 Base Gemini (conocimiento interno)"],
        index=0,
        label_visibility="collapsed"
    )

    expandir_tema = st.checkbox("🔎 Buscar sinónimos y términos relacionados", value=True,
                                  help="Gemini expandirá tu tema a variantes semánticas (sinónimos, inglés/español, subcampos) antes de buscar")

    st.markdown("---")
    st.markdown("### 🌎 Región prioritaria")
    reg_norteamerica = st.checkbox("Norteamérica (EE.UU. / Canadá)", value=True)
    reg_europa       = st.checkbox("Europa Occidental, Central y Norte", value=True)
    reg_latam        = st.checkbox("América Latina y el Caribe", value=True)

    st.markdown("---")
    st.markdown("### 📚 Tipo de membresía")
    tipo_institucional = st.checkbox("Institucional", value=True,
        help="Universidades completas, consorcios, bibliotecas — acceso a nivel organización")
    tipo_corporativa = st.checkbox("Corporativa", value=True,
        help="Empresas, clústeres tecnológicos, partners industriales")
    tipo_academica = st.checkbox("Académica / Docente", value=True,
        help="Profesores, investigadores titulares, directores de tesis")
    tipo_estudiantil = st.checkbox("Estudiantil", value=True,
        help="Pregrado y posgrado con dominio .edu")
    tipo_individual = st.checkbox("Individual Profesional", value=True,
        help="Profesionistas independientes, freelance, certificación personal")
    tipo_asociado = st.checkbox("Asociado / Afiliado", value=True,
        help="Nivel de entrada más bajo, comunidad o alumni")

    st.markdown("---")
    st.markdown("### ⭐ Palabras clave prioritarias")
    st.caption("Si aparecen en el nombre, Gemini les dará prioridad")
    keyword_selections = {}

    custom_keyword = st.text_input(
        "Agregar palabra clave (solo esta búsqueda)",
        placeholder="Ej: Association, League, Alliance, Society, Charter, Royal College, Organization...",
        key="custom_keyword_input"
    )

    st.markdown("---")
    st.markdown("### 🧭 Modalidad Búsqueda")
    modalidad_costo = st.radio(
        "Modalidad",
        options=["Fase de Negociación", "Vinculación directa"],
        index=1,
        label_visibility="collapsed"
    )

    # Las 8 condiciones quedan fijas en el prompt, ya no son seleccionables en UI
    if modalidad_costo == "Vinculación directa":
        cond_total = cond_edu = cond_grant = cond_sandbox = True
        cond_alumni = cond_cert = cond_beta = cond_prueba = True
    else:
        cond_total = cond_edu = cond_grant = cond_sandbox = False
        cond_alumni = cond_cert = cond_beta = cond_prueba = False




# ── Helpers de filtros ───────────────────────────────────────────────────────
def build_filters_summary():
    regiones, tipos, condiciones, keywords = [], [], [], []
    if reg_norteamerica: regiones.append("Norteamérica (EE.UU. y Canadá)")
    if reg_europa:       regiones.append("Europa Occidental, Central y Norte")
    if reg_latam:        regiones.append("América Latina y el Caribe")

    if tipo_institucional: tipos.append("Institucional")
    if tipo_corporativa:   tipos.append("Corporativa")
    if tipo_academica:     tipos.append("Académica/Docente")
    if tipo_estudiantil:   tipos.append("Estudiantil")
    if tipo_individual:    tipos.append("Individual Profesional")
    if tipo_asociado:      tipos.append("Asociado/Afiliado")

    if cond_total:   condiciones.append("Gratuidad total")
    if cond_edu:     condiciones.append("Student Tier (.edu o dominio universitario)")
    if cond_grant:   condiciones.append("Grant-Based Free Tier")
    if cond_sandbox: condiciones.append("Institutional Sandbox")
    if cond_alumni:  condiciones.append("Año de Gracia / Alumni Launchpad")
    if cond_cert:    condiciones.append("Gratuidad por certificación o cursos completados")
    if cond_beta:    condiciones.append("Acceso Beta Tester")
    if cond_prueba:  condiciones.append("Prueba institucional de 12 meses")

    # Métodos de acceso siempre activos (ya no configurables en UI)
    accesos = [
        "SSO Institucional o IP Whitelisting",
        "Dominio .edu verificado",
        "Invitación o aprobación previa",
        "Cuenta personal con validación de perfil"
    ]

    for kw, selected in keyword_selections.items():
        if selected:
            keywords.append(kw)

    if custom_keyword and custom_keyword.strip():
        keywords.append(custom_keyword.strip())

    return regiones, tipos, condiciones, accesos, keywords


def build_prompt(topic, regiones, tipos, condiciones, accesos, keywords, n, excluidas=None, use_search=True, expandir_tema=True, buscar_pago=False, pdf_topics=None):
    keyword_block = (
        ", ".join(keywords) if keywords
        else "ninguna palabra clave especificada"
    )

    search_instruction = (
        "Usa búsqueda activa en internet (Google Search grounding) para encontrar plataformas reales y verificar que las URLs existen actualmente."
        if use_search else
        "Usa tu conocimiento interno sin buscar en internet en tiempo real. Marca url_verificada como false si no estás seguro de que la URL exista actualmente."
    )

    expansion_instruction = ""
    if expandir_tema and topic:
        expansion_instruction = f"""
EXPANSIÓN SEMÁNTICA DEL TEMA (paso obligatorio antes de buscar):
Antes de buscar, genera mentalmente una lista de 5 a 8 sinónimos, términos relacionados, variantes en inglés/español y subcampos del tema "{topic}". Por ejemplo, si el tema es "gestión de referencias bibliográficas", variantes incluirían: reference management, citation management, bibliografía, manejo de citas, gestores bibliográficos, EndNote alternatives, software de citación académica.
Usa TODAS esas variantes como consultas de búsqueda adicionales, no solo el término literal escrito por el usuario. Esto es crítico: muchas organizaciones usan terminología distinta a la del usuario aunque cubran el mismo dominio.
"""

    pdf_topics_block = ""
    if pdf_topics:
        ordered_list = chr(10).join(f"{i+1}. {t}" for i, t in enumerate(pdf_topics))
        pdf_topics_block = f"""
TEMAS DETECTADOS EN EL DOCUMENTO CARGADO (orden de prioridad, el primero pesa más):
{ordered_list}
Usa estos temas como ejes principales de búsqueda, en el orden indicado. El tema #1 debe dominar la mayoría de los resultados; los siguientes se usan para diversificar si el primero no genera suficientes coincidencias. Trátalos con el mismo peso que el tema de búsqueda escrito por el usuario, combinándolos si es necesario.
"""

    if buscar_pago:
        objetivo_costo = "que requieran PAGO (membresías de pago estándar, sin necesidad de que sean gratuitas)"
        regla_costo = "1. Incluye membresías de pago. No es necesario que sean gratuitas ni que tengan condiciones especiales."
        condiciones_block = "No aplica filtro de gratuidad — incluye cualquier estructura de precio."
    else:
        objetivo_costo = "GRATUITAS (sin costo alguno, periodo mínimo 12 meses o 1 año)"
        regla_costo = "1. Solo membresías COMPLETAMENTE GRATUITAS. Excluye cualquier opción de pago."
        condiciones_block = (
            chr(10).join(f"- {c}" for c in condiciones) if condiciones else "- Cualquier condición de gratuidad"
        )

    return f"""Eres un agente especializado en inteligencia de recursos académicos y profesionales. Tu función es identificar, evaluar y catalogar plataformas web que ofrecen membresías, suscripciones o accesos institucionales para el sector educativo y de investigación.

MODO DE BÚSQUEDA: {search_instruction}

TEMA DE BÚSQUEDA: {topic if topic else "herramientas y recursos académicos generales"}
{expansion_instruction}{pdf_topics_block}
OBJETIVO: Encuentra exactamente {n} páginas web que ofrezcan membresías académicas o institucionales {objetivo_costo} para el tema indicado y sus variantes semánticas.

REGIONES PRIORITARIAS (busca SOLO en estas):
{chr(10).join(f"- {r}" for r in regiones) if regiones else "- Sin filtro regional"}

TIPOS DE MEMBRESÍA A BUSCAR:
{chr(10).join(f"- {t}" for t in tipos) if tipos else "- Todos los tipos"}

CONDICIONES ACEPTABLES:
{condiciones_block}

MÉTODOS DE ACCESO ACEPTABLES (siempre se consideran todos):
- SSO Institucional o IP Whitelisting
- Dominio .edu verificado
- Invitación o aprobación previa
- Cuenta personal con validación de perfil

PALABRAS CLAVE DE PRIORIDAD: Si el nombre oficial de la organización contiene alguna de estas palabras, dale prioridad y aumenta su puntuación (suelen ser asociaciones, ligas, alianzas o colegios profesionales con mayor probabilidad de ofrecer membresías institucionales):
{keyword_block}

REGLAS OBLIGATORIAS:
{regla_costo}
2. El período de acceso debe ser de 12 meses o 1 año como mínimo.
3. Excluye plataformas de: Asia Oriental/Pacífico, Asia del Sur/Central, África, Medio Oriente.
4. Excluye organismos multilaterales: ONU, UNESCO, FMI, BM, OCDE, OMS, OEA y equivalentes.
5. Prioriza plataformas con poca visibilidad sobre las ampliamente conocidas, y prioriza aquellas cuyo nombre contenga las palabras clave indicadas.
6. Nunca inventes URLs. Si no puedes verificar un dato, usa "Sin verificar".
7. No incluyas contenido detrás de login previo que no sea descubrible públicamente.
8. Ordena los resultados de mayor a menor puntuación (5 a 1).
9. NUNCA repitas las siguientes membresías que ya han sido obtenidas previamente (lista de exclusión):
{{EXCLUSION_BLOCK}}

FORMATO DE RESPUESTA: Responde ÚNICAMENTE con un array JSON válido, sin texto adicional, sin bloques de código markdown. El array debe tener exactamente {n} objetos:

[
  {{
    "nombre": "Nombre oficial de la organización",
    "descripcion": "2-3 líneas explicando de qué trata la asociación/organización y a qué se dedica",
    "url": "https://url-directa-a-pagina-de-membresia.com",
    "precio": "Gratis / Gratis con condición específica",
    "condicion_gratuidad": "Tipo específico de gratuidad que aplica",
    "metodo_acceso": "Método de acceso requerido",
    "beneficios": ["Beneficio 1", "Beneficio 2", "Beneficio 3", "Beneficio 4"],
    "link_membresia": "URL directa al formulario o página para solicitar la membresía",
    "correo_contacto": "Correo de contacto si está disponible públicamente, o 'No disponible'",
    "tipo_membresia": "Institucional|Corporativa|Académica/Docente|Estudiantil|Individual Profesional|Asociado/Afiliado",
    "region": "Norteamérica|Europa|América Latina|Global",
    "puntuacion": 4,
    "url_verificada": true,
    "duracion": "12 meses / 1 año / Indefinida mientras seas estudiante",
    "contiene_palabra_clave": true
  }}
]""".replace("{{EXCLUSION_BLOCK}}", ("\n".join(f"- {e}" for e in excluidas) if excluidas else "ninguna por ahora"))


def clean_json_string(raw):
    raw = re.sub(r"```json", "", raw)
    raw = re.sub(r"```", "", raw).strip()
    raw = re.sub(r",\s*([}\]])", r"\1", raw)
    raw = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", raw)
    m = re.search(r"\[.*\]", raw, re.DOTALL)
    if m:
        raw = m.group()
    return raw


def parse_json_results(raw):
    """Intenta parsear el JSON; si está truncado, rescata los objetos completos."""
    raw = clean_json_string(raw)

    try:
        results = json.loads(raw)
        return results if isinstance(results, list) else [results]
    except json.JSONDecodeError:
        pass

    for suffix in ["}]", "}]}]", "]"]:
        try:
            results = json.loads(raw + suffix)
            return results if isinstance(results, list) else [results]
        except json.JSONDecodeError:
            pass

    results = []
    depth = 0
    start = None
    for i, ch in enumerate(raw):
        if ch == "{":
            if depth == 0:
                start = i
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0 and start is not None:
                fragment = raw[start:i+1]
                fragment = clean_json_string(fragment)
                try:
                    results.append(json.loads(fragment))
                except json.JSONDecodeError:
                    pass
                start = None
    if results:
        return results

    raise ValueError(
        f"No se pudo parsear la respuesta de Gemini.\n"
        f"Intenta reducir el número de resultados o vuelve a buscar.\n\n"
        f"Fragmento recibido:\n{raw[:300]}"
    )


def run_search(topic, regiones, tipos, condiciones, accesos, keywords, n, use_search=True, expandir_tema=True, buscar_pago=False, pdf_topics=None):
    excluidas = load_existing_memberships()
    prompt = build_prompt(topic, regiones, tipos, condiciones, accesos, keywords, n, excluidas, use_search, expandir_tema, buscar_pago, pdf_topics)
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel("gemini-2.5-flash")

    needed_tokens = max(8192, n * 600 + 2000)

    gen_config = genai.types.GenerationConfig(
        temperature=0.3,
        max_output_tokens=needed_tokens
    )

    tools = None
    if use_search:
        tools = "google_search_retrieval"

    try:
        if use_search:
            response = model.generate_content(
                prompt,
                generation_config=gen_config,
                tools=[{"google_search": {}}]
            )
        else:
            response = model.generate_content(prompt, generation_config=gen_config)
    except Exception:
        # Fallback sin tools si el modelo/SDK no soporta google_search en esta versión
        response = model.generate_content(prompt, generation_config=gen_config)

    raw = response.text.strip()

    finish = None
    try:
        finish = response.candidates[0].finish_reason.name
    except Exception:
        pass

    results = parse_json_results(raw)

    if finish == "MAX_TOKENS" and len(results) < n:
        st.session_state["truncated_warning"] = (
            f"⚠️ Gemini truncó la respuesta ({len(results)} de {n} resultados recuperados). "
            "Intenta pedir menos resultados o pulsa 🔄 Nuevos resultados."
        )
    else:
        st.session_state.pop("truncated_warning", None)

    return results


def render_stars(score):
    return f'<span class="stars">{"★" * score}</span><span class="stars stars-empty">{"☆" * (5 - score)}</span>'


def render_type_badge(tipo):
    t = (tipo or "").lower()
    mapping = {
        "institucional": "badge-type-institucional",
        "corporativa": "badge-type-corporativa",
        "académica/docente": "badge-type-academica",
        "academica/docente": "badge-type-academica",
        "académica": "badge-type-academica",
        "academica": "badge-type-academica",
        "estudiantil": "badge-type-estudiantil",
        "individual profesional": "badge-type-individual",
        "individual": "badge-type-individual",
        "asociado/afiliado": "badge-type-asociado",
        "asociado": "badge-type-asociado",
        "afiliado": "badge-type-asociado",
    }
    cls = mapping.get(t, "badge-type-institucional")
    return f'<span class="badge {cls}">{tipo}</span>'


def render_results(results):
    if "decisions" not in st.session_state:
        st.session_state.decisions = {}

    for i, r in enumerate(results):
        score    = r.get("puntuacion", 3)
        verified = r.get("url_verificada", None)
        if verified is True:
            ver_html = '<span class="verified-badge verified-ok">✓ URL verificada</span>'
        elif verified is False:
            ver_html = '<span class="verified-badge verified-no">⚠ Sin verificar</span>'
        else:
            ver_html = '<span class="verified-badge verified-unk">? Por verificar</span>'

        priority_html = ""
        if r.get("contiene_palabra_clave"):
            priority_html = '<span class="badge badge-priority">⭐ Prioridad</span>'

        url            = r.get("url", "#")
        nombre         = r.get("nombre", "Sin nombre")
        descripcion    = r.get("descripcion", "")
        precio         = r.get("precio", "Gratis")
        link_membresia = r.get("link_membresia", url)
        correo_default = r.get("correo_contacto", "")
        bene_html      = "".join(f"<li>{b}</li>" for b in r.get("beneficios", []))

        decision = st.session_state.decisions.get(i, {"estatus": "Pendiente", "correo": correo_default})

        card_class = "result-card"
        if decision["estatus"] == "Aprobado":
            card_class += " approved"
        elif decision["estatus"] == "Rechazado":
            card_class += " rejected"

        st.markdown(f"""
        <div class="{card_class}">
            <div class="card-title">#{i+1} — {nombre} {ver_html} {priority_html}</div>
            <div class="card-url"><a href="{url}" target="_blank">{url}</a></div>
            <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:0.6rem;">
                {render_stars(score)}
                {render_type_badge(r.get("tipo_membresia",""))}
                <span class="badge badge-region">{r.get("region","")}</span>
                <span class="badge badge-access">🕐 {r.get("duracion","12 meses")}</span>
            </div>
            <div class="card-section-title">De qué trata</div>
            <div class="card-text">{descripcion}</div>
            <div class="card-section-title">Precio</div>
            <div class="card-text">{precio}</div>
            <div class="card-section-title">Condición de gratuidad</div>
            <div class="card-text">{r.get("condicion_gratuidad","")}</div>
            <div class="card-section-title">Método de acceso</div>
            <div class="card-text">{r.get("metodo_acceso","")}</div>
            <div class="card-section-title">Beneficios principales</div>
            <div class="card-text"><ul style="margin:0;padding-left:1.2rem;">{bene_html}</ul></div>
            <div class="card-section-title">Solicitar membresía</div>
            <div class="card-text"><a href="{link_membresia}" target="_blank">{link_membresia}</a></div>
        </div>
        """, unsafe_allow_html=True)

        col_a, col_b, col_c = st.columns([1.3, 2, 1])
        with col_a:
            estatus = st.radio(
                f"Decisión #{i+1}",
                options=["Pendiente", "Aprobado", "Rechazado"],
                index=["Pendiente", "Aprobado", "Rechazado"].index(decision["estatus"]),
                key=f"estatus_{i}",
                horizontal=True,
                label_visibility="collapsed"
            )
        with col_b:
            correo = st.text_input(
                f"Correo de contacto #{i+1}",
                value=decision.get("correo", correo_default),
                key=f"correo_{i}",
                placeholder="correo@institucion.edu",
                label_visibility="collapsed"
            )
        with col_c:
            st.write("")

        st.session_state.decisions[i] = {"estatus": estatus, "correo": correo}
        st.markdown("<div style='margin-bottom:0.8rem;'></div>", unsafe_allow_html=True)


def build_excel(results, topic):
    wb = Workbook()
    ws = wb.active
    ws.title = "Membresías Gratuitas"

    COLOR_HEADER_BG = "0D2B4E"
    COLOR_HEADER_FG = "E8B84B"
    COLOR_SUBHEADER  = "C8973A"
    COLOR_SUBHEADER_FG = "FFFFFF"
    COLOR_ROW_ALT   = "F4F6F9"
    COLOR_BORDER    = "C8D0DC"
    COLOR_SCORE_5   = "1A5232"
    COLOR_SCORE_4   = "3F7D5C"
    COLOR_SCORE_3   = "C8973A"
    COLOR_SCORE_2   = "B5651D"
    COLOR_SCORE_1   = "7B1E1E"

    thin = Side(style="thin", color=COLOR_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = f"ARIA Membresías — {topic or 'Recursos Académicos'}"
    title_cell.font = Font(name="Arial", bold=True, size=14, color=COLOR_HEADER_FG)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:M2")
    date_cell = ws["A2"]
    date_cell.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total resultados: {len(results)}"
    date_cell.font = Font(name="Arial", size=10, color="6B7A8D")
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    headers = ["#", "Nombre", "Descripción", "URL", "Precio", "Condición de gratuidad",
               "Método de acceso", "Beneficios", "Link Membresía", "Correo Contacto",
               "Tipo", "Región", "Puntuación"]
    col_widths = [4, 26, 40, 32, 20, 28, 26, 45, 32, 24, 16, 16, 12]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color=COLOR_SUBHEADER_FG)
        cell.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[3].height = 22

    score_colors = {5: COLOR_SCORE_5, 4: COLOR_SCORE_4, 3: COLOR_SCORE_3, 2: COLOR_SCORE_2, 1: COLOR_SCORE_1}

    for row_idx, r in enumerate(results, start=4):
        alt = row_idx % 2 == 0
        row_fill = PatternFill("solid", fgColor=COLOR_ROW_ALT) if alt else PatternFill("solid", fgColor="FFFFFF")

        beneficios_txt = "\n".join(f"• {b}" for b in r.get("beneficios", []))
        score = r.get("puntuacion", 3)
        stars = "★" * score + "☆" * (5 - score)

        values = [
            row_idx - 3,
            r.get("nombre", ""),
            r.get("descripcion", ""),
            r.get("url", ""),
            r.get("precio", "Gratis"),
            r.get("condicion_gratuidad", ""),
            r.get("metodo_acceso", ""),
            beneficios_txt,
            r.get("link_membresia", r.get("url", "")),
            r.get("correo_contacto", ""),
            r.get("tipo_membresia", ""),
            r.get("region", ""),
            stars,
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True,
                                       horizontal="center" if col_idx in [1, 11, 12, 13] else "left")

            if col_idx == 13:
                sc = score_colors.get(score, "888888")
                cell.font = Font(name="Arial", size=10, bold=True, color=sc)
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = max(60, 15 * len(r.get("beneficios", [])))

    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("Filtros aplicados")
    ws2["A1"] = "Parámetro"
    ws2["B1"] = "Valor"
    for cell in [ws2["A1"], ws2["B1"]]:
        cell.font = Font(name="Arial", bold=True, size=10, color=COLOR_SUBHEADER_FG)
        cell.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
        cell.border = border
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 60

    meta = [
        ("Tema de búsqueda", topic or "Recursos académicos generales"),
        ("Fecha de generación", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Total resultados", len(results)),
        ("Regiones", ", ".join([r for r in ["Norteamérica" if reg_norteamerica else "", "Europa" if reg_europa else "", "América Latina" if reg_latam else ""] if r])),
        ("Tipos de membresía", ", ".join([t for t, v in [
            ("Institucional", tipo_institucional), ("Corporativa", tipo_corporativa),
            ("Académica/Docente", tipo_academica), ("Estudiantil", tipo_estudiantil),
            ("Individual Profesional", tipo_individual), ("Asociado/Afiliado", tipo_asociado)] if v])),
        ("Condiciones de gratuidad", ", ".join([c for c, v in [("Gratuidad total", cond_total), ("Student Tier", cond_edu), ("Grant-Based", cond_grant), ("Sandbox", cond_sandbox), ("Alumni Launchpad", cond_alumni), ("Certificación", cond_cert), ("Beta Tester", cond_beta), ("Prueba 12m", cond_prueba)] if v])),
    ]
    for r_idx, (k, v) in enumerate(meta, start=2):
        ws2.cell(row=r_idx, column=1, value=k).font = Font(name="Arial", size=9, bold=True)
        ws2.cell(row=r_idx, column=2, value=str(v)).font = Font(name="Arial", size=9)
        ws2.cell(row=r_idx, column=1).border = border
        ws2.cell(row=r_idx, column=2).border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Mascota animada del investigador ─────────────────────────────────────────
def get_researcher_mascot_svg(state="idle"):
    """state: 'idle' o 'searching' — controla la animación vía clase CSS."""
    cls = "mascot-searching" if state == "searching" else "mascot-idle"
    return f'''<svg class="{cls}" width="56" height="56" viewBox="0 0 56 56" xmlns="http://www.w3.org/2000/svg">
<g class="mascot-body">
  <circle cx="28" cy="20" r="9" fill="#E8B84B" stroke="#0D2B4E" stroke-width="1.2"/>
  <path d="M20 18 Q28 10 36 18" fill="none" stroke="#0D2B4E" stroke-width="1.5" stroke-linecap="round"/>
  <rect x="19" y="29" width="18" height="17" rx="6" fill="#1A3F6F" stroke="#0D2B4E" stroke-width="1.2"/>
  <rect x="23" y="29" width="10" height="6" rx="2" fill="#2E5FA3"/>
  <circle cx="24" cy="20" r="1.4" fill="#0D2B4E"/>
  <circle cx="32" cy="20" r="1.4" fill="#0D2B4E"/>
</g>
<g class="mag-glass" style="transform-origin: 40px 30px;">
  <circle cx="40" cy="28" r="6.5" fill="rgba(232,184,75,0.18)" stroke="#E8B84B" stroke-width="2"/>
  <line x1="44.5" y1="32.5" x2="49" y2="37" stroke="#E8B84B" stroke-width="2.4" stroke-linecap="round"/>
</g>
</svg>'''


# ── Extracción de temas clave desde dossier o temario PDF ────────────────────
def extract_topics_from_pdf(pdf_file, max_chars=6000):
    """
    Extrae texto del PDF (priorizando secciones de temario/plan de estudios si existen)
    y usa una llamada a Gemini para identificar 3-6 temas clave de búsqueda.
    Funciona tanto con dossiers/CVs personales como con planes de estudio o folletos
    institucionales que listan asignaturas o áreas de conocimiento.
    Devuelve (lista_de_temas, None) en éxito, o (None, mensaje_error) en fallo.
    """
    try:
        reader = PdfReader(pdf_file)
        all_pages_text = []
        for page in reader.pages:
            all_pages_text.append(page.extract_text() or "")

        full_doc = " ".join(all_pages_text)
        full_doc = re.sub(r"\s+", " ", full_doc).strip()

        if len(full_doc) < 50:
            return None, "El PDF no contiene texto extraíble (puede ser un escaneo de imagen sin OCR)."

        keywords_priority = ["asignatura", "plan de estudio", "temario", "contenido",
                              "índice", "syllabus", "módulo", "unidad", "curso"]
        priority_text = ""
        other_text = ""
        for page_text in all_pages_text:
            low = page_text.lower()
            if any(k in low for k in keywords_priority):
                priority_text += " " + page_text
            else:
                other_text += " " + page_text

        combined = re.sub(r"\s+", " ", priority_text).strip()
        if len(combined) < max_chars:
            filler = re.sub(r"\s+", " ", other_text).strip()
            combined = (combined + " " + filler)[:max_chars]
        else:
            combined = combined[:max_chars]

        if not api_key:
            return None, "No hay GEMINI_API_KEY configurada en los Secrets."

        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-2.5-flash")

        mini_prompt = f"""Lee el siguiente fragmento de un documento (puede ser un dossier/CV de una persona, un plan de estudios, un temario de asignaturas, o un folleto de un programa académico).

Identifica entre 3 y 6 TEMAS O ÁREAS DE CONOCIMIENTO clave que mejor representen el contenido técnico o académico del documento. Estos temas se usarán para buscar membresías o recursos académicos relacionados, así que deben ser específicos y útiles como término de búsqueda (ej: "bases de datos avanzadas", "ciberseguridad", "inteligencia artificial", "gestión de proyectos tecnológicos"), no genéricos como "educación" o "tecnología".

Responde ÚNICAMENTE con un array JSON de strings, sin texto adicional ni bloques de código. Ejemplo de formato: ["tema 1", "tema 2", "tema 3"]

Fragmento del documento:
{combined}

Temas clave:"""

        try:
            response = model.generate_content(
                mini_prompt,
                generation_config=genai.types.GenerationConfig(temperature=0.2, max_output_tokens=800)
            )
        except Exception as api_err:
            return None, f"Error al llamar a Gemini: {str(api_err)}"

        try:
            raw = response.text.strip()
        except Exception:
            finish_reason = None
            try:
                finish_reason = response.candidates[0].finish_reason
            except Exception:
                pass
            return None, f"Gemini no devolvió texto utilizable (finish_reason={finish_reason}). Puede ser un bloqueo de seguridad o respuesta vacía."

        raw_clean = re.sub(r"```json", "", raw)
        raw_clean = re.sub(r"```", "", raw_clean).strip()

        try:
            topics = json.loads(raw_clean)
            if isinstance(topics, list) and topics:
                return [str(t).strip() for t in topics if str(t).strip()], None
        except json.JSONDecodeError:
            match = re.search(r"\[.*\]", raw_clean, re.DOTALL)
            if match:
                try:
                    topics = json.loads(match.group())
                    if isinstance(topics, list) and topics:
                        return [str(t).strip() for t in topics if str(t).strip()], None
                except json.JSONDecodeError:
                    pass

            # JSON truncado: rescatar todos los strings completos "..." encontrados
            rescued = re.findall(r'"([^"]+)"', raw_clean)
            if rescued:
                return [s.strip() for s in rescued if s.strip()], None

        return None, f"Gemini respondió pero no en formato JSON esperado. Respuesta cruda: {raw[:300]}"

    except Exception as ex:
        return None, f"Error inesperado procesando el PDF: {str(ex)}"


# ── Main UI ───────────────────────────────────────────────────────────────────
if "is_searching" not in st.session_state:
    st.session_state.is_searching = False

mascot_state = "searching" if st.session_state.is_searching else "idle"
mascot_label = "Buscando..." if st.session_state.is_searching else "Listo para buscar"
mascot_svg = get_researcher_mascot_svg(mascot_state)

header_placeholder = st.empty()
header_placeholder.markdown(f"""
<div class="inst-header">
    <div class="inst-header-left">
        <div class="inst-logo">🎓</div>
        <div>
            <div class="inst-title">ARIA <span>Membresías</span></div>
            <div class="inst-subtitle">Alliance Recognition &amp; Intelligence Architecture</div>
        </div>
    </div>
    <div class="inst-mascot-box">
        {mascot_svg}
        <div class="inst-mascot-label">{mascot_label}</div>
    </div>
</div>
""", unsafe_allow_html=True)

topic_col, btn_col = st.columns([4, 1])
with topic_col:
    topic = st.text_input(
        "Tema o dominio de búsqueda",
        value=st.session_state.get("topic_from_pdf", ""),
        placeholder="Ej: software de análisis estadístico, gestión de referencias, diseño curricular...",
        key="topic_input"
    )
with btn_col:
    st.markdown("<br>", unsafe_allow_html=True)
    buscar = st.button("🔍 Buscar", use_container_width=True)

# ── Session state ─────────────────────────────────────────────────────────────
if "results" not in st.session_state:
    st.session_state.results = None
if "last_topic" not in st.session_state:
    st.session_state.last_topic = ""
if "decisions" not in st.session_state:
    st.session_state.decisions = {}
if "mem_pdf_topics_hidden" not in st.session_state:
    st.session_state.mem_pdf_topics_hidden = None
if "mem_trigger_pdf_search" not in st.session_state:
    st.session_state.mem_trigger_pdf_search = False

with st.expander("📄 O cargar un dossier, temario o plan de estudios en PDF para iniciar la búsqueda automáticamente"):
    pdf_file = st.file_uploader(
        "Sube un dossier académico, CV, temario de asignaturas o folleto de programa (PDF)",
        type=["pdf"],
        key="mem_pdf_dossier_uploader",
        help="Se extrae solo el texto (sin imágenes) para detectar temas de búsqueda — no se envía el PDF completo a Gemini"
    )
    if pdf_file is not None:
        if st.button("🔎 Analizar y buscar", key="mem_detect_topic_btn", use_container_width=True):
            with st.spinner("Analizando el documento..."):
                detected_topics, error_msg = extract_topics_from_pdf(pdf_file)
            if detected_topics:
                st.session_state.mem_pdf_topics_hidden = detected_topics
                st.session_state.mem_trigger_pdf_search = True
                st.rerun()
            else:
                st.error(f"No se pudieron identificar temas: {error_msg}")



def do_search():
    if not api_key:
        st.error("Configura GEMINI_API_KEY en los Secrets de Streamlit Cloud.")
        return
    regiones, tipos, condiciones, accesos, keywords = build_filters_summary()
    if not regiones:
        st.warning("Selecciona al menos una región en el panel lateral.")
        return
    if modalidad_costo == "Vinculación directa" and not condiciones:
        st.warning("Selecciona al menos una condición de costo preferencial.")
        return

    use_search = modo_busqueda.startswith("🔴")

    st.session_state.is_searching = True
    header_placeholder.markdown(f"""
    <div class="inst-header">
        <div class="inst-header-left">
            <div class="inst-logo">🎓</div>
            <div>
                <div class="inst-title">ARIA <span>Membresías</span></div>
                <div class="inst-subtitle">Alliance Recognition &amp; Intelligence Architecture</div>
            </div>
        </div>
        <div class="inst-mascot-box">
            {get_researcher_mascot_svg("searching")}
            <div class="inst-mascot-label">Buscando...</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    try:
        pdf_topics = st.session_state.get("mem_pdf_topics_hidden")
        results = run_search(topic, regiones, tipos, condiciones, accesos, keywords, MAX_RESULTS, use_search, expandir_tema, modalidad_costo == "Fase de Negociación", pdf_topics)
        st.session_state.results = results
        st.session_state.last_topic = topic
        st.session_state.decisions = {}
    except Exception as e:
        st.error(f"Error: {str(e)}")
    finally:
        st.session_state.is_searching = False
        header_placeholder.markdown(f"""
        <div class="inst-header">
            <div class="inst-header-left">
                <div class="inst-logo">🎓</div>
                <div>
                    <div class="inst-title">ARIA <span>Membresías</span></div>
                    <div class="inst-subtitle">Alliance Recognition &amp; Intelligence Architecture</div>
                </div>
            </div>
            <div class="inst-mascot-box">
                {get_researcher_mascot_svg("idle")}
                <div class="inst-mascot-label">Búsqueda completada</div>
            </div>
        </div>
        """, unsafe_allow_html=True)


if buscar or st.session_state.get("mem_trigger_pdf_search"):
    st.session_state.mem_trigger_pdf_search = False
    do_search()

# ── Mostrar resultados ────────────────────────────────────────────────────────
if st.session_state.results:
    results = st.session_state.results

    col_status, col_regen, col_excel, col_save = st.columns([2.5, 1.2, 1.2, 1.2])
    with col_status:
        st.markdown(
            f'<div class="status-bar">✦ {len(results)} membresías encontradas</div>',
            unsafe_allow_html=True
        )
    with col_regen:
        if st.button("🔄 Nuevos resultados", use_container_width=True):
            do_search()
            st.rerun()
    with col_excel:
        excel_buf = build_excel(results, st.session_state.last_topic)
        fname = f"ARIA_Membresias_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            label="📥 Descargar Excel",
            data=excel_buf,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with col_save:
        if st.button("💾 Guardar en Sheet", use_container_width=True):
            with st.spinner("Guardando en Google Sheets..."):
                ok, info = save_to_sheets(results, st.session_state.decisions)
            if ok:
                added = info.get("added", 0)
                updated = info.get("updated", 0)
                if added or updated:
                    st.success(f"✓ {added} nuevas · {updated} actualizadas")
                else:
                    st.info("Sin cambios para guardar")
                st.cache_data.clear()
            else:
                st.error(f"Error al guardar: {info}")

    if st.session_state.get("truncated_warning"):
        st.warning(st.session_state["truncated_warning"])

    render_results(results)

else:
    st.markdown("""
    <div class="empty-state">
        <span class="empty-icon">🎓</span>
        Configura los filtros en el panel lateral y escribe un tema para comenzar.<br>
        <span style="font-size:0.82rem; color:#9E9B90;">Los filtros siempre están visibles — ajústalos antes de cada búsqueda.</span>
    </div>
    """, unsafe_allow_html=True)
